import os
import lxml.etree as et
from openpyxl import load_workbook


# holds the keyboard entry value at start up
class CharVar(object):
    pass


def set_def_directory():
    # Start with either (I) Import or (E) Export
    message = input("(I) Import or (E) Export an XML file:  ")
    # Normalise path for all OS version
    folder = '~/Desktop/Mars_TEAM'
    desktop = os.path.normpath(folder)
    p = os.path.expanduser(desktop)
    # navigate to the default folder path of the user profile/ desktop folder.
    # folder = 'Desktop/Mars_TEAM'
    # p = os.path.expanduser('~/'+folder)
    dir_path = p
    # print(dir_path)
    os.chdir(dir_path)
    # place holder for value at start
    # print(message)
    prompt_for_execution_option(message)


def prompt_for_execution_option(message):

    xml_file = input("Enter XML file name: ")

    excel_file = input("Enter EXCEL file name: ")

    if message == "I":
        # Import need 1 excel 1 xml file
        # get file name from user message.e Mexican template copy.xml
        print("is the file name correct?", xml_file)
        # ask for the file name:
        read_xml(xml_file, excel_file)
    elif message == "E":
        # Export need need 1 xml 1 excel file
        # get file name from user message.e template.xml
        # write_file = input("Enter the XML file name to Export to: ")
        # ask for the file name:
        read_excel_file(excel_file, xml_file)
    else:
        print("Please use either I (Import) or E (Export) to Use this application?")
        exit()


def read_excel_file(excel_file, xml_file):
    # work on the excel file to read all entry from Column A
    reference_uri = []
    reference_identifier = []
    book = load_workbook(excel_file)
    sheet = book[xml_file]
    for rows in range(2, 25):
        value = sheet.cell(row=rows, column=1)
        if sheet.cell(row=rows, column=1).value is not None:
            reference_uri.append(value.value)
            # print(value.value)
        else:
            break

    for rows2 in range(26, 36):
        value2 = sheet.cell(row=rows2, column=1)
        if sheet.cell(row=rows2, column=1).value is not None:
            reference_identifier.append(value2.value)
            # print(value2.value)
        else:
            break

    parse_xml(xml_file, reference_uri, reference_identifier)
    # print("end of read excel function")


def parse_xml(xml_file, reference_uri, reference_identifier):
    # look for the entries in the XML file:
    print(xml_file)
    tree2 = et.parse(xml_file)
    root2 = tree2.findall(".//referenceIdentifier")

    parse_root(reference_uri, xml_file)

    print("\n")

    parse_root2(root2, reference_identifier, tree2)

    print("End of - UPDATING YOUR, XML FILE")


def parse_root(reference_uri, xml_file):
    i = 0
    tree = et.parse(xml_file)
    root = tree.getroot()
    for item in root.findall('.//employee'):
        print(len(root))
        if 'employee' in item.tag:
            print(len(item))
            while len(reference_uri) != 0:
                msg = reference_uri.pop(i)
                new_element = et.SubElement(item, "MyNewEmployee")
                new_reference_uri = et.SubElement(new_element, "Text")
                new_reference_uri.text = msg
                print(new_reference_uri, new_reference_uri.text)
    # tree.append(root)
    et.register_namespace('artwork_content', 'urn:gs1:ecom:artwork_content:xsd:3')
    et.register_namespace('sh', 'http://www.unece.org/cefact/namespaces/StandardBusinessDocumentHeader')
    et.register_namespace('xsi', 'http://www.w3.org/2001/XMLSchema-instance')
    # writing out to a new file since the writing to the same file does not seem to work.
    tree.write("NEW_FILE2.xml", xml_declaration=True, encoding='utf-8', method="xml")
    return print("write to referenceUri is completed")


def parse_root2(root2, reference_identifier, tree2):
    i = 0
    print(i)

    for elem in root2:
        msg2 = reference_identifier.pop(i)
        elem.text = msg2
        print(msg2)
        tree2.write('NEW_FILE2.xml')

    return print("write to referenceUri is completed")


def compare_file_names(xml_file, excel_file):
    wb = load_workbook(excel_file)
    sheet_names = wb.get_sheet_names
    base_file = os.path.basename(xml_file)
    file_name = os.path.splitext(base_file)[0]
    for sheet in sheet_names:
        # print(sheet)
        if sheet == file_name:
            print('found match' + sheet)
            return sheet
        else:
            print('no match found')


def read_xml(xml_file, excel_file):
    x = xml_file
    # holds value from xml file
    item = []
    item2 = []
    # read xml file
    tree = et.parse(x)
    tree2 = et.parse(x)
    root3 = tree.findall(".//referenceURI")
    root4 = tree2.findall(".//referenceIdentifier")

    for items in root3:
        item.append(items)

    for items2 in root4:
        item2.append(items2)

    write_excel(item, item2, xml_file, excel_file)


def write_excel(item, item2, xml_file, excel_file):
    xml = item
    xml2 = item2
    f = excel_file
    wb = load_workbook(f)
    base_file = os.path.basename(xml_file)
    wb.create_sheet(title=base_file)
    current_sheet = wb[base_file]
    # add the 2 column header to the sheet
    header = current_sheet.cell(row=1, column=1)
    header.value = "Path to Image file on server / Do not Change layout."
    # writing to the excel file here:
    i = 0
    row = 2
    # print(len(xml))
    while len(xml) != 0:
        va = xml.pop(i)
        cell_ref = current_sheet.cell(row=row, column=1)
        cell_ref.value = va.text
        row = row + 1
        # print(va.text)

    # add the 2 column header to the sheet
    header2 = current_sheet.cell(row=25, column=1)  # last row +1
    header2.value = "Reference Identifier"
    t = 0
    rows = 26  # or row +2
    # print(len(xml2))
    while len(xml2) != 0:
        val = xml2.pop(t)
        cell_ref = current_sheet.cell(row=rows, column=1)
        cell_ref.value = val.text
        rows = rows + 1
        print(val.text)

    # end of read xml file
    # write to the excel file
    wb.save(f)
    print('writing to file is completed ', wb.sheetnames)
    # end to write ops

# set_def_directory() the starting point


if __name__ == "__main__":
    set_def_directory()

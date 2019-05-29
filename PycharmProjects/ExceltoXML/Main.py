
import os
import lxml.etree as et
from openpyxl import load_workbook


# change to the directory where all files saved
# check if the application is running on OS X or Windows
def get_os_version():
    
    # Normalise path for all OS version
    folder = '~/Desktop/Mars_TEAM'
    desktop = os.path.normpath(folder)
    p = os.path.expanduser(desktop)
    # navigate to the default folder path of the user profile/ desktop folder.
    # folder = 'Desktop/Mars_TEAM'
    # p = os.path.expanduser('~/'+folder)
    dir_path = p
    # print(dir_path)
    os.chdir(dir_path)
    # place holder for value at start
    message2 = input("Import or Export an XML File:  ")
    print(message2)
    return message2


def read_excel_file(f_path2):
    # work on the excel file to read all entry from Column A
    file = f_path2
    reference_uri = []
    reference_identifier = []
    book = load_workbook(file)
    sheet = book[xml_file]
    row_count = sheet.max_row
    for rows in range(2, row_count+1):
        value = sheet.cell(row=rows, column=1)
        reference_uri.append(value.value)

    for rows2 in range(26, row_count + 1):
        value2 = sheet.cell(row=rows2, column=1)
        reference_identifier.append(value2.value)

    parse_xml(file_to_parse, reference_uri, reference_identifier)
    # print("end of read excel function")


def parse_xml(file_to_parse, reference_uri, reference_identifier):
    # look for the entries in the XML file:
    print(file_to_parse)
    tree = et.parse(file_to_parse)
    tree2 = et.parse(file_to_parse)
    root = tree.findall(".//referenceURI")
    root2 = tree2.findall(".//referenceIdentifier")

    i = 0
    for elem in root:
        msg = reference_uri.pop()
        elem.text = msg
        # print(elem.text)
        tree.write(file_to_parse)

    for element in root2:
        msg2 = reference_identifier.pop()
        element.text = msg2
        # print(element.text)
        tree2.write(file_to_parse)

    print("End of - LOOKUP IN XML FILE")


def compare_file_names(file, file_to_parse):
    wb = load_workbook(file)
    sheet_names = wb.get_sheet_names
    base_file = os.path.basename(file_to_parse)
    file_name = os.path.splitext(base_file)[0]
    for sheet in sheet_names:
        # print(sheet)
        if sheet == file_name:
            print('found match' + sheet)
            return sheet
        else:
            print('no match found')


def read_xml(xml_file_read):
    x = xml_file_read
    # holds value from xml file
    item = []
    item2 = []
    # read xml file
    tree = et.parse(x)
    tree2 = et.parse(x)
    root3 = tree.findall(".//referenceURI")
    root4 = tree2.findall(".//referenceIdentifier")

    for items in root3:
        item.append(items)

    for items2 in root4:
        item2.append(items2)

    write_excel(item, item2)


def write_excel(item, item2):
    xml = item
    xml2 = item2
    f = file_to_write
    wb = load_workbook(f)
    base_file = os.path.basename(xml_file_read)
    wb.create_sheet(title=base_file)
    current_sheet = wb[base_file]
    # add the 2 column header to the sheet
    header = current_sheet.cell(row=1, column=1)
    header.value = "Path to Image file on server / Do not Change layout."
    # writing to the excel file here:
    i = 0
    row = 2
    # print(len(xml))
    while len(xml) != 0:
        va = xml.pop(i)
        cell_ref = current_sheet.cell(row=row, column=1)
        cell_ref.value = va.text
        row = row + 1
        # print(va.text)

    # add the 2 column header to the sheet
    header2 = current_sheet.cell(row=25, column=1)  # last row +1
    header2.value = "Reference Identifier"
    t = 0
    rows = 26  # or row +2
    # print(len(xml2))
    while len(xml2) != 0:
        val = xml2.pop(t)
        cell_ref = current_sheet.cell(row=rows, column=1)
        cell_ref.value = val.text
        rows = rows + 1
        print(val.text)

    # end of read xml file
    # write to the excel file
    wb.save(f)
    print('writing to file is completed ', wb.sheetnames)
    # end to write ops


if __name__ == "__main__":
    message = get_os_version()
    if message == "I":
        # message to display
        write_file = "Enter XML file name to Import: "
        # get file name from user message.e Mexican template copy.xml
        xml_file_read = input(write_file)
        print("is the file name correct?", xml_file_read)
        # ask for the file name:
        ask_for_input2 = "Enter EXCEL file name to Use: "
        file_to_write = input(ask_for_input2)
        read_xml(xml_file_read)
    elif message == "E":
        # XML TO Update
        ask_for_input = "Enter the XML file name to Export to: "
        # get file name from user message.e template.xml
        xml_file = input(ask_for_input)
        print("is the file name correct?  ", xml_file)
        # Select the file to work with XMLTemplate.xlsx
        file_to_parse = xml_file
        # ask for the file name:
        ask_for_input2 = "Enter EXCEL file name to Use: "
        excel_file = input(ask_for_input2)
        f_path2 = excel_file
        read_excel_file(f_path2)
    else:
        print("Please use either I (Import) or E (Export) to Use this application?")
        exit()






